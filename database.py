# database.py
"""
Handles all JSON and Excel manipulation for the Ballistic Tests Database application.
Centralizes database operations (reading/writing JSON, processing Excel files) for modularity.
"""

import os
import glob
import shutil
import json
import numpy as np
from datetime import datetime, timedelta
from openpyxl import load_workbook
from tkinter import messagebox
import config
from utils import clean_value, parse_date


def create_daily_backup(json_file):
    """Create a daily backup of the JSON database if it doesn't already exist."""
    try:
        if not os.path.exists(json_file):
            print(f"Backup skipped: {json_file} not found.")
            return

        base_dir = os.path.dirname(json_file) or "."
        backup_dir = os.path.join(base_dir, "Backup")

        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
            print(f"Created Backup directory: {backup_dir}")

        current_date = datetime.now().strftime("%Y%m%d")
        backup_pattern = os.path.join(backup_dir, f"Data_{current_date}_*.json")
        existing_backups = glob.glob(backup_pattern)

        if existing_backups:
            print(f"Backup already exists for {current_date}: {existing_backups[0]}")
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"Data_{timestamp}.json"
        backup_path = os.path.join(backup_dir, backup_filename)

        shutil.copy2(json_file, backup_path)
        print(f"Backup created: {backup_path}")
    except Exception as e:
        print(f"Error creating backup: {str(e)}")


def process_orders(orders_input, json_file, excel_folder):
    """Process Excel files for given orders and update the JSON database."""
    try:
        if not orders_input:
            return False, "Please enter the order numbers."

        orders = [order.strip() for order in orders_input.split(",")]

        data = {}
        if os.path.exists(json_file):
            with open(json_file, "r", encoding="utf-8") as f:
                data = json.load(f)

        files_to_process = []
        for order in orders:
            for file_name in os.listdir(excel_folder):
                if file_name.startswith(order) and file_name.endswith(".xlsx"):
                    files_to_process.append(os.path.join(excel_folder, file_name))

        if not files_to_process:
            return False, "No Excel files found for the provided orders."

        for file in files_to_process:
            process_excel(file, data)

        with open(json_file, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        return True, "Excel files processed and JSON updated!"
    except Exception as e:
        return False, f"Error processing orders: {str(e)}"


def process_excel(file_path, data):
    """Process an Excel file and update the data dictionary."""
    wb = load_workbook(file_path, data_only=True)
    current_version = None
    current_order = None
    for sheet_name in wb.sheetnames:
        if "minus" in sheet_name.lower():
            temp_type = "LT"
        elif "rt" in sheet_name.lower():
            temp_type = "RT"
        elif "plus" in sheet_name.lower():
            temp_type = "HT"
        else:
            continue

        if "datenblatt" in sheet_name.lower():
            current_version, current_order = process_datenblatt(
                wb[sheet_name], temp_type, data
            )
        elif "grafik" in sheet_name.lower() and current_version and current_order:
            process_grafik(
                wb[sheet_name], temp_type, data, current_version, current_order
            )


def process_datenblatt(sheet, temp_type, data):
    """Extract metadata and test data from a Datenblatt sheet."""
    inflator_type = clean_value(sheet["U1"].value)
    version = "V" + inflator_type.split("V")[-1]

    test_order = clean_value(sheet["J4"].value)
    production_order = clean_value(sheet["J3"].value)
    propellant_lot_number = clean_value(sheet["S3"].value)
    test_date = parse_date(sheet["C4"].value)
    temperature_c = clean_value(sheet["C10"].value)

    if version not in data:
        data[version] = {}

    if test_order not in data[version]:
        data[version][test_order] = {
            "metadata": {
                "production_order": production_order,
                "propellant_lot_number": propellant_lot_number,
                "test_date": test_date,
            },
            "temperatures": {},
        }

    if temp_type not in data[version][test_order]["temperatures"]:
        data[version][test_order]["temperatures"][temp_type] = {
            "temperature_c": float(temperature_c) if temperature_c else None,
            "tests": [],
        }

    tests = []
    seen_tests = set()
    for row in sheet.iter_rows(min_row=10, values_only=True):
        if row[0] and str(row[0]).strip().isdigit():
            test_no = clean_value(row[0])
            inflator_no = clean_value(row[1])
            if test_no and inflator_no and test_no not in seen_tests:
                tests.append({"test_no": int(test_no), "inflator_no": int(inflator_no)})
                seen_tests.add(test_no)

    data[version][test_order]["temperatures"][temp_type]["tests"] = tests
    return version, test_order


def process_grafik(sheet, temp_type, data, current_version, current_order):
    """Extract pressure data and limits from a Grafik sheet."""
    valid_columns = []
    limits = {"maximums": {}, "minimums": {}}
    for col in range(config.MIN_COLUMN, config.MAX_COLUMN):
        min_val = clean_value(sheet.cell(row=config.MIN_LIMIT_ROW, column=col).value)
        max_val = clean_value(sheet.cell(row=config.MAX_LIMIT_ROW, column=col).value)
        if min_val or max_val:
            valid_columns.append(col)
            ms = col - 2
            if min_val is not None:
                try:
                    limits["minimums"][str(ms)] = float(min_val)
                except ValueError:
                    pass
            if max_val is not None:
                try:
                    limits["maximums"][str(ms)] = float(max_val)
                except ValueError:
                    pass

    inflator_nos = [
        test["inflator_no"]
        for test in data[current_version][current_order]["temperatures"][temp_type][
            "tests"
        ]
    ]

    pressure_data = []
    blank_line_count = 0
    row_idx = config.PRESSURE_DATA_START_ROW

    for inflator_no in inflator_nos:
        is_blank = True
        pressures = {}
        for col in valid_columns:
            pressure = clean_value(sheet.cell(row=row_idx, column=col).value)
            if pressure is not None:
                try:
                    pressures[str(col - 2)] = float(pressure)
                    is_blank = False
                except ValueError:
                    continue

        if is_blank:
            blank_line_count += 1
            if blank_line_count >= 2:
                break
        else:
            blank_line_count = 0
            if pressures:
                pressure_data.append(
                    {"inflator_no": inflator_no, "pressures": pressures}
                )

        row_idx += 1

    data[current_version][current_order]["temperatures"][temp_type][
        "pressure_data"
    ] = pressure_data
    data[current_version][current_order]["temperatures"][temp_type]["limits"] = limits


def remove_orders(orders_input, json_file):
    """Remove specified orders from the JSON database."""
    try:
        if not orders_input:
            return False, "Please enter the order numbers."

        orders_to_remove = [order.strip() for order in orders_input.split(",")]

        if not os.path.exists(json_file):
            return False, "No database found."

        with open(json_file, "r", encoding="utf-8") as f:
            data = json.load(f)

        removed = []
        for version in list(data.keys()):
            for order in list(data[version].keys()):
                if order in orders_to_remove:
                    del data[version][order]
                    removed.append(order)
                    if not data[version]:
                        del data[version]

        with open(json_file, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        if removed:
            msg = f"Orders removed successfully: {', '.join(removed)}"
            return True, msg
        else:
            return False, "No matching orders found."
    except Exception as e:
        return False, f"Error removing orders: {str(e)}"


def get_orders_list(json_file, selected_version, start_date, end_date):
    """Retrieve a filtered list of orders from the JSON database."""
    if not os.path.exists(json_file):
        return [], [], "Database not found."

    try:
        with open(json_file, "r", encoding="utf-8") as f:
            data = json.load(f)

        versions = sorted(data.keys())
        orders_list = []
        for version, orders in data.items():
            if (
                selected_version
                and selected_version.lower() != "all"
                and version != selected_version
            ):
                continue
            for order, details in orders.items():
                test_date = details["metadata"].get("test_date", "0000-00-00")
                if start_date:
                    try:
                        test_date_obj = datetime.strptime(test_date, "%Y-%m-%d").date()
                        if not (start_date <= test_date_obj <= end_date):
                            continue
                    except ValueError:
                        continue
                orders_list.append((version, order, test_date))

        def parse_date_safe(date_str):
            try:
                return datetime.strptime(date_str, "%Y-%m-%d")
            except Exception:
                return datetime.min

        orders_list.sort(key=lambda x: parse_date_safe(x[2]), reverse=True)
        return orders_list, versions, None
    except Exception as e:
        return [], [], f"Error loading orders: {str(e)}"


def get_metadata(json_file, version, order):
    """Retrieve metadata for a specific order and version."""
    if not os.path.exists(json_file):
        return None, "Database not found."

    try:
        with open(json_file, "r", encoding="utf-8") as f:
            data = json.load(f)

        if version not in data or order not in data[version]:
            return None, "Order not found in the database."

        metadata = data[version][order].get("metadata", {})
        temperatures = data[version][order].get("temperatures", {})
        return {"metadata": metadata, "temperatures": temperatures}, None
    except Exception as e:
        return None, f"Error retrieving metadata: {str(e)}"


def get_workplace_data(json_file, selected_orders):
    """Retrieve data for selected orders to send to the workplace."""
    try:
        if not selected_orders:
            return [], "No orders selected."

        with open(json_file, "r", encoding="utf-8") as f:
            data = json.load(f)

        versions = set(version for version, _ in selected_orders)
        if len(versions) > 1:
            return [], "Select tests from only one version."

        new_workplace_data = []
        duplicates_skipped = 0
        existing_keys = set()  # Track existing entries to avoid duplicates

        for version, order in selected_orders:
            if version in data and order in data[version]:
                details = data[version][order]
                metadata = details.get("metadata", {})
                test_date = metadata.get("test_date", "0000-00-00")
                temperatures = details.get("temperatures", {})
                for temp_type in ["RT", "LT", "HT"]:
                    if temp_type not in temperatures:
                        continue
                    temp_data = temperatures[temp_type]
                    temperature_c = temp_data.get("temperature_c", "N/A")
                    tests = temp_data.get("tests", [])
                    pressure_data = temp_data.get("pressure_data", [])
                    pressure_map = {
                        item["inflator_no"]: item["pressures"] for item in pressure_data
                    }
                    for test in tests:
                        test_no = test.get("test_no", "N/A")
                        inflator_no = test.get("inflator_no", "N/A")
                        key = (test_no, inflator_no, temp_type, version, order)
                        if key in existing_keys:
                            duplicates_skipped += 1
                            continue
                        new_workplace_data.append(
                            {
                                "test_no": test_no,
                                "inflator_no": inflator_no,
                                "temperature_c": temperature_c,
                                "type": temp_type,
                                "version": version,
                                "order": order,
                                "test_date": test_date,
                                "pressures": pressure_map.get(inflator_no, {}),
                            }
                        )
                        existing_keys.add(key)

        def parse_date_safe(date_str):
            try:
                return datetime.strptime(date_str, "%Y-%m-%d")
            except:
                return datetime(1900, 1, 1)

        new_workplace_data.sort(
            key=lambda x: parse_date_safe(x.get("test_date", "1900-01-01")),
            reverse=True,
        )

        msg = f"Added {len(new_workplace_data)} records."
        if duplicates_skipped > 0:
            msg += f"\n{duplicates_skipped} duplicate test(s) were ignored."
        return new_workplace_data, msg
    except Exception as e:
        return [], f"Error retrieving workplace data: {str(e)}"
